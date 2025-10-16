#!/usr/bin/env python3
"""
Excel file management for ModelScore application.
Handles multi-tab Excel files with timestamped filenames.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelManager:
    """Manages Excel file operations with multiple tabs."""
    
    def __init__(self, output_dir: Path, filename_prefix: str = "modelscore"):
        """Initialize the Excel manager.
        
        Args:
            output_dir: Directory where Excel files will be saved
            filename_prefix: Prefix for the Excel filename
        """
        self.output_dir = output_dir
        self.filename_prefix = filename_prefix
        self.logger = logging.getLogger(__name__)
        
        # Generate timestamped filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.filename = f"{filename_prefix}_{timestamp}.xlsx"
        self.file_path = output_dir / self.filename
        
        # Initialize workbook
        self.workbook = Workbook()
        # Remove default sheet
        self.workbook.remove(self.workbook.active)
        
        self.logger.info(f"Excel manager initialized. Output file: {self.file_path}")
    
    def _sanitize_tab_name(self, name: str) -> str:
        """Sanitize tab name to comply with Excel naming rules.
        
        Args:
            name: Original tab name
            
        Returns:
            Sanitized tab name
        """
        # Excel tab names have restrictions:
        # - Max 31 characters
        # - Cannot contain: / \ ? * [ ]
        # - Cannot be empty
        
        # Replace invalid characters
        invalid_chars = ['/', '\\', '?', '*', '[', ']']
        sanitized = name
        for char in invalid_chars:
            sanitized = sanitized.replace(char, '_')
        
        # Truncate if too long
        if len(sanitized) > 31:
            sanitized = sanitized[:31]
        
        # Ensure not empty
        if not sanitized:
            sanitized = "Sheet"
        
        return sanitized
    
    def create_tab(self, tab_name: str, data: List[Dict[str, Any]], 
                   headers: Optional[List[str]] = None) -> None:
        """Create a new tab with data in the Excel file.
        
        Args:
            tab_name: Name of the tab (will be sanitized)
            data: List of dictionaries containing the data
            headers: Optional list of column headers
        """
        sanitized_tab_name = self._sanitize_tab_name(tab_name)
        
        # Check if tab already exists and append number if needed
        original_name = sanitized_tab_name
        counter = 1
        while sanitized_tab_name in [sheet.title for sheet in self.workbook.worksheets]:
            sanitized_tab_name = f"{original_name}_{counter}"
            counter += 1
        
        # Create new worksheet
        worksheet = self.workbook.create_sheet(title=sanitized_tab_name)
        
        if not data:
            self.logger.warning(f"No data provided for tab: {sanitized_tab_name}")
            return
        
        # Convert data to DataFrame for easier handling
        df = pd.DataFrame(data)
        
        # Write headers if provided
        if headers:
            worksheet.append(headers)
        else:
            # Use DataFrame columns as headers
            worksheet.append(list(df.columns))
        
        # Write data rows
        for _, row in df.iterrows():
            worksheet.append(list(row))
        
        self.logger.info(f"Created tab '{sanitized_tab_name}' with {len(data)} rows")
    
    def create_tab_from_key_value_pairs(self, tab_name: str, 
                                       data: Dict[str, Any]) -> None:
        """Create a tab from key-value pairs (like model info).
        
        Args:
            tab_name: Name of the tab
            data: Dictionary of key-value pairs
        """
        # Convert to list of dictionaries format
        rows = [{"Key": key, "Value": str(value)} for key, value in data.items()]
        self.create_tab(tab_name, rows, ["Key", "Value"])
    
    def create_tab_from_csv_data(self, tab_name: str, 
                                data: List[Dict[str, Any]]) -> None:
        """Create a tab from CSV-like data (list of dictionaries).
        
        Args:
            tab_name: Name of the tab
            data: List of dictionaries with consistent keys
        """
        self.create_tab(tab_name, data)
    
    def save(self) -> Path:
        """Save the Excel file to disk.
        
        Returns:
            Path to the saved Excel file
        """
        try:
            self.workbook.save(self.file_path)
            self.logger.info(f"Excel file saved: {self.file_path}")
            return self.file_path
        except Exception as e:
            self.logger.error(f"Error saving Excel file: {e}")
            raise
    
    def get_file_path(self) -> Path:
        """Get the file path for the Excel file.
        
        Returns:
            Path to the Excel file
        """
        return self.file_path
    
    def get_filename(self) -> str:
        """Get the filename of the Excel file.
        
        Returns:
            Excel filename
        """
        return self.filename
